"""
Report Generation Utilities
Handles PDF and Excel export for various reports
"""

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ReportGenerator:
    """Base class for report generation"""
    
    def __init__(self, title, company_name="AssetHub"):
        self.title = title
        self.company_name = company_name
        self.generated_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # ==================== PDF GENERATION ====================
    
    def generate_pdf(self, data, columns, filename="report.pdf"):
        """Generate a PDF report"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Header
        elements.append(Paragraph(self.company_name, title_style))
        elements.append(Paragraph(self.title, styles['Heading2']))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Generated: {self.generated_date}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Table data
        table_data = [columns] + data
        
        # Create table
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4361ee')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Body styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    # ==================== EXCEL GENERATION ====================
    
    def generate_excel(self, data, columns, filename="report.xlsx"):
        """Generate an Excel report"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = self.title[:31]  # Excel limits sheet names to 31 chars
        
        # Styling
        header_fill = PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        sheet.merge_cells('A1:' + get_column_letter(len(columns)) + '1')
        title_cell = sheet['A1']
        title_cell.value = f"{self.company_name} - {self.title}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = center_align
        
        # Generated date
        sheet.merge_cells('A2:' + get_column_letter(len(columns)) + '2')
        date_cell = sheet['A2']
        date_cell.value = f"Generated: {self.generated_date}"
        date_cell.font = Font(italic=True, size=10)
        date_cell.alignment = center_align
        
        # Headers (row 4)
        for col_num, column_title in enumerate(columns, 1):
            cell = sheet.cell(row=4, column=col_num)
            cell.value = column_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Data
        for row_num, row_data in enumerate(data, 5):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Auto-adjust column widths
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column].width = adjusted_width
        
        # Save to buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer


class AssetReportGenerator(ReportGenerator):
    """Generate asset-specific reports"""
    
    def generate_asset_inventory_report(self, assets, format='pdf'):
        """Generate complete asset inventory report"""
        columns = ['Asset ID', 'Type', 'Brand', 'Model', 'Serial', 'Status', 'Department', 'Cost (₵)']
        
        data = []
        for asset in assets:
            data.append([
                asset.get('id', 'N/A'),
                asset.get('type', 'N/A'),
                asset.get('brand', 'N/A'),
                asset.get('model', 'N/A'),
                asset.get('serial', 'N/A'),
                asset.get('status', 'N/A'),
                asset.get('department', 'N/A'),
                f"₵{asset.get('cost', 0):.2f}"
            ])
        
        if format == 'pdf':
            return self.generate_pdf(data, columns)
        else:
            return self.generate_excel(data, columns)
    
    def generate_depreciation_report(self, depreciation_data, format='pdf'):
        """Generate depreciation report"""
        columns = ['Asset ID', 'Type', 'Brand', 'Department', 'Purchase Cost', 'Current Value', 'Depreciation', 'Age (Years)']
        
        data = []
        for item in depreciation_data:
            data.append([
                item.get('asset_id', 'N/A'),
                item.get('type', 'N/A'),
                item.get('brand', 'N/A'),
                item.get('department', 'N/A'),
                f"₵{item.get('purchase_cost', 0):.2f}",
                f"₵{item.get('current_value', 0):.2f}",
                f"₵{item.get('depreciation', 0):.2f}",
                f"{item.get('years_old', 0):.1f}"
            ])
        
        if format == 'pdf':
            return self.generate_pdf(data, columns)
        else:
            return self.generate_excel(data, columns)


class MaintenanceReportGenerator(ReportGenerator):
    """Generate maintenance-specific reports"""
    
    def generate_maintenance_report(self, maintenance_records, format='pdf'):
        """Generate maintenance cost report"""
        columns = ['Asset ID', 'Date', 'Activity', 'Cost (₵)', 'Technician', 'Notes']
        
        data = []
        for record in maintenance_records:
            data.append([
                record.get('asset_id', 'N/A'),
                record.get('date', 'N/A'),
                record.get('activity', 'N/A'),
                f"₵{record.get('cost', 0):.2f}",
                record.get('technician', 'N/A'),
                record.get('notes', '')[:50] + '...' if len(record.get('notes', '')) > 50 else record.get('notes', '')
            ])
        
        if format == 'pdf':
            return self.generate_pdf(data, columns)
        else:
            return self.generate_excel(data, columns)


class DepartmentReportGenerator(ReportGenerator):
    """Generate department-specific reports"""
    
    def generate_department_summary(self, departments_data, format='pdf'):
        """Generate department summary report"""
        columns = ['Department', 'Code', 'Total Assets', 'Active Assets', 'Total Cost (₵)', 'Utilization %']
        
        data = []
        for dept in departments_data:
            data.append([
                dept.get('name', 'N/A'),
                dept.get('code', 'N/A'),
                dept.get('total_assets', 0),
                dept.get('active_assets', 0),
                f"₵{dept.get('total_cost', 0):.2f}",
                f"{dept.get('utilization_rate', 0):.1f}%"
            ])
        
        if format == 'pdf':
            return self.generate_pdf(data, columns)
        else:
            return self.generate_excel(data, columns)